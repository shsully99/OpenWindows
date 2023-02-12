import shutil
#import os

import datetime
from openpyxl import Workbook
from openpyxl import load_workbook


def DownloadFile (selectedelements, facadedetails, gsVolume, gsRev, ElementSRI, BaseDir):
    print ("def DownloadFile (selectedelements, facadedetails, gsVolume, gsRev, ElementSRI):")
    # Copy xls file to a unique time stamped copy
    
    src = BaseDir + "/" + '_InternalCalc.xlsx'
    #dst = r'C:\Users\Administrator.SHAREPOINTSKY\Desktop\Newfolder\name.txt'

    dst=BaseDir + "/" + datetime.datetime.now().strftime("sample_files/%Y-%m-%d-%H-%M-%S-%f"+".xlsx")
    print ("source " + src)

    print ("destination " + dst)

    try:
        shutil.copyfile(src, dst)
        print("File copied successfully.")
 
    # If source and destination are same
    except shutil.SameFileError:
        print("Source and destination represents the same file.")
        return "", "Source and destination represents the same file."


    # If there is any permission issue
    except PermissionError:
        print("Permission denied.")
        return "", "Permission denied."
 
    # For other errors
    except:
        print("Error occurred while copying file.")
        return "", "Error occurred while copying file."

    # Open our unique copy and automate it using

    wb = load_workbook(dst)

    #' Set up basic info 

    ws = wb["Sheet1"]

    ws['L5'].value=gsRev
    ws['L6'].value=gsVolume

    ' Loop  through and assign data to the excel datasheets'

    i=7
    for facade in facadedetails:
        if facade["FacadeSpectra"] != "":
            if facade["Metric"] == "Laeq16":
                ws.cell(row = i, column = 2).value = "Daytime LAeq,16h dB(A)"
            elif facade["Metric"] == "Laeq8":
                ws.cell(row = i, column = 2).value = "Night-time LAeq,8h dB(A)"
            elif facade["Metric"] == "LamaxV":
                ws.cell(row = i, column = 2).value = "Ventilation LAFmax dB(A)"
            elif facade["Metric"] == "LamaxO":
                ws.cell(row = i, column = 2).value = "Overheating LAFmax dB(A)"

            sIncident = [float(x) for x in facade["FacadeSpectra"].rstrip(';').split("-")]

            ws.cell(row = i, column = 5).value = sIncident[0]
            ws.cell(row = i, column = 6).value = sIncident[1]
            ws.cell(row = i, column = 7).value = sIncident[2]
            ws.cell(row = i, column = 8).value = sIncident[3]
            ws.cell(row = i, column = 9).value = sIncident[4]

            i=i+1

    i = 16

    for selement in selectedelements: 

        ourElement = ElementSRI.query.filter(ElementSRI.UniqueID == selement["ElementID"] ).first()
        print (ourElement.Description)
        ws.cell(row = i, column = 2).value = ourElement.Description
        ws.cell(row = i, column = 3).value = selement["Quantity"]
        ws.cell(row = i, column = 4).value = selement["FacadeDifference"]
        ws.cell(row = i, column = 5).value = selement["Hz125"]
        ws.cell(row = i, column = 6).value = selement["Hz250"]
        ws.cell(row = i, column = 7).value = selement["Hz500"]
        ws.cell(row = i, column = 8).value = selement["Hz1000"]
        ws.cell(row = i, column = 9).value = selement["Hz2000"]

        ws.cell(row = i, column = 10).value = ourElement.Metric
        ws.cell(row = i, column = 11).value = selement["State"]
        i=i+1

    wb.save(dst)
    wb.close()
    
    return "Success", dst

    #os.rename(r'C:\Users\Administrator.SHAREPOINTSKY\Desktop\Work\name.txt',r'C:\Users\Administrator.SHAREPOINTSKY\Desktop\Newfolder\details.txt' )

